import yaml
import logging
from typing import List, Tuple
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import os
import json
import time
import pathlib
import pandas as pd
from labs import scripts
from openpyxl import load_workbook
import urllib3


class BundleDetailsCollector:
    def __init__(self):
        self.driver = None
        urllib3.disable_warnings(urllib3.exceptions.HTTPWarning)

        with open("./configs/configurations.yaml", "r") as file:
            self.configurations = yaml.safe_load(file)
        self.mergedOutputFolderPath: str = self.configurations["mergedOutputFolderPath"]
        self.undallocatedBundlesFolder: str = self.configurations[
            "undallocatedBundlesFolder"
        ]
        self.cookies_path = "./configs/cookies.json"

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )

    @staticmethod
    def log_message(message: str, level: int = logging.INFO) -> None:
        logging.log(level, message)

    def load_webdriver(self) -> None:
        if self.driver is None:
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # Run in headless mode (optional)
            # chrome_options.add_argument("--no-sandbox")
            # chrome_options.add_argument("--disable-dev-shm-usage")

            # self.service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(options=chrome_options)

        self.driver.get("https://examerp.keralauniversity.ac.in/")

    def close_webdriver(self) -> None:
        if self.driver:
            self.driver.quit()
            self.driver = None
        if self.service:
            self.service.stop()
            self.service = None

    def fetch_csrftoken(self) -> str:
        cookies = self.driver.get_cookies()
        return next(
            (cookie["value"] for cookie in cookies if cookie["name"] == "csrftoken"),
            None,
        )

    def attempt_login(self) -> Tuple[bool, str]:
        try:
            self.log_message("Attempting login", logging.INFO)
            captcha_link = self.driver.find_element(
                By.CLASS_NAME, "captcha"
            ).get_attribute("src")
            self.log_message(f"captcha_link {captcha_link}", logging.INFO)
            return False, captcha_link
        except Exception as e:
            self.log_message(f"Attempt login failed: {e}", logging.ERROR)
            return False, ""

    def manual_login(self, username: str, password: str, capcha_text: str) -> bool:
        try:
            self.driver.find_element(By.ID, "username").send_keys(username)
            self.driver.find_element(By.ID, "password").send_keys(password)

            capcha_field = self.driver.find_element(By.ID, "id_captcha_1")
            capcha_field.clear()

            self.log_message(
                f"capcha_text {capcha_text} {username} {password}", logging.INFO
            )
            capcha_field.send_keys(capcha_text)
            self.driver.find_element(By.CLASS_NAME, "btn-primary").click()
            self.create_cookies()
            return True
        except Exception as e:
            self.log_message(f"Manual login failed: {e}", logging.ERROR)
            return False

    def check_cookies(self) -> Tuple[bool, str]:
        if not self.driver:
            self.load_webdriver()
        self.log_message("Checking Cookies", logging.INFO)

        if not os.path.exists(self.cookies_path):
            self.log_message("No cookies file found", logging.WARNING)
            return self.attempt_login()

        try:
            with open(self.cookies_path, "r") as f:
                cookies = json.load(f)

            if cookies[0]["expiry"] <= int(time.time()):
                self.log_message("Cookies have expired", logging.WARNING)
                return self.attempt_login()
            self.add_cookies(cookies=cookies)
            self.log_message("Cookies are valid. Redirecting...", logging.INFO)
            return True, ""

        except Exception as e:
            self.log_message(f"Error reading cookies: {e}", logging.ERROR)
            return self.attempt_login()

    def add_cookies(self, cookies: List[dict]) -> None:
        for cookie in cookies:
            self.driver.add_cookie(cookie)
        self.driver.get(
            "https://examerp.keralauniversity.ac.in/cd-unit/qpcode-wise-bundle-list"
        )

    def create_cookies(self) -> None:
        with open(self.cookies_path, "w") as f:
            json.dump(self.driver.get_cookies(), f, indent=4)
        self.log_message("Cookies captured. Redirecting...", logging.INFO)
        self.driver.get(
            "https://examerp.keralauniversity.ac.in/cd-unit/qpcode-wise-bundle-list"
        )

    def custom_qp_range(self, qp_codes: List[str], exam_name: str, log_server) -> None:
        pathlib.Path(self.mergedOutputFolderPath).mkdir(parents=True, exist_ok=True)
        path = pathlib.Path(self.undallocatedBundlesFolder + exam_name)
        path.mkdir(parents=True, exist_ok=True)

        try:
            self.grab_bundle_details_using_script(
                qp_codes=qp_codes,
                path=str(path) + "/",
                exam_name=exam_name,
                log_server=log_server,
            )
        except Exception as e:
            error_message = f"Error executing Script in fun custom_qp_range: {e}"
            log_server(error_message, "ERROR")
            self.log_message(error_message, logging.ERROR)

    def auto_qp_series_generator(
        self, series: str, range_start: int, range_end: int, exam_name: str, log_server
    ) -> None:
        qp_codes = [f"{series} {i}" for i in range(range_start, range_end + 1)]
        pathlib.Path(self.mergedOutputFolderPath).mkdir(parents=True, exist_ok=True)
        path = pathlib.Path(self.undallocatedBundlesFolder + exam_name)
        path.mkdir(parents=True, exist_ok=True)
        try:
            self.grab_bundle_details_using_script(
                qp_codes=qp_codes,
                path=str(path) + "/",
                exam_name=exam_name,
                log_server=log_server,
            )
        except Exception as e:
            log_server("Error fetching bundle details", "ERROR")
            self.log_message(f"Error executing Script: {e}", logging.ERROR)

    def grab_bundle_details_using_script(
        self, qp_codes: List[str], path: str, exam_name: str, log_server
    ) -> None:
        headers = [
            "Sl.No.",
            "Bundle Code",
            "AS Count",
            "Course Name",
            "District",
            "Camp",
            "Status",
        ]

        for qp_code in qp_codes:
            data = []
            self.log_message(f"Fetching details of: {qp_code}", logging.INFO)
            log_server(f"Fetching details of: {qp_code}", "INFO")

            res = self.driver.execute_script(
                scripts.get_bundles_list(
                    csrftoken=self.fetch_csrftoken(),
                    qp_code=qp_code,
                )
            )
            try:
                if res["message"] == "success":
                    self.log_message(f"Saving Details of: {qp_code}", logging.INFO)
                    for bundle_details in res["data"]["bundleList"][0]:
                        status = (
                            "Allocated To Camp"
                            if bundle_details["status"] == "DELIVERED UNIVERSITY"
                            else bundle_details["status"]
                        )
                        data.append(
                            [
                                "",
                                bundle_details["bundleCode"],
                                bundle_details["totalCount"],
                                bundle_details["courseName"],
                                bundle_details["district"],
                                bundle_details["camp"],
                                status,
                            ]
                        )
                    df = pd.DataFrame(data, columns=headers)
                    file_name = f"{exam_name} {qp_code}"
                    final_file_name = f"{path}{file_name}.xlsx"
                    df.to_excel(final_file_name, index=False)
                else:
                    self.log_message(res["message"], logging.WARNING)
                    log_server(res["message"], "WARNING")
            except Exception as e:
                self.log_message(
                    f"Error fetched result {res}| Error: {e}", logging.ERROR
                )
                log_server(f"Error fetched result {res}| Error: {e}", "ERROR")

        output_file = f"{self.mergedOutputFolderPath}{exam_name}_combined.xlsx"
        self.merge_excel_files(path, output_file, exam_name, log_server)

    def merge_excel_files(
        self, folder_path: str, output_file: str, exam_name: str, log_server
    ) -> None:
        merged_data = pd.DataFrame()
        for file in os.listdir(folder_path):
            if file.startswith(exam_name) and file.endswith(".xlsx"):
                file_path = os.path.join(folder_path, file)
                data = pd.read_excel(file_path)
                merged_data = pd.concat([merged_data, data], ignore_index=True)
        merged_data.sort_values(by="Bundle Code", ascending=True, inplace=True)

        self.style_merged_table(merged_data, output_file)
        self.log_message(f"Merged data saved to {output_file}", logging.INFO)
        log_server(f"Merged data saved to {output_file}", "INFO")
        log_server("Merging all bundle details", "INFO")

    def style_merged_table(self, merged_data, output_file):
        df = merged_data
        with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
            workbook = writer.book
            df.drop(df.columns[[0]], axis=1, inplace=True)

            for camp in df["Camp"].unique():
                try:
                    new_df = df[df["Camp"] == camp].reset_index()
                    new_df = new_df.rename(columns={"index": "Sl.No"})
                    new_df["Sl.No"] = new_df.index + 1
                    new_df.to_excel(
                        writer, sheet_name=camp, index=False, startrow=1, header=False
                    )

                    worksheet = writer.sheets[camp]
                    self.apply_styles(worksheet, new_df, workbook)
                except Exception:
                    new_df = df[df["Camp"].isnull()].reset_index()
                    new_df = new_df.rename(columns={"index": "Sl.No"})
                    new_df["Sl.No"] = new_df.index + 1
                    new_df.to_excel(
                        writer,
                        sheet_name="Generated",
                        index=False,
                        startrow=1,
                        header=False,
                    )

                    worksheet = writer.sheets["Generated"]
                    self.apply_styles(worksheet, new_df, workbook)

    def apply_styles(self, worksheet, df, workbook):
        align_format = workbook.add_format(
            {"font_size": 11, "align": "center", "valign": "vcenter", "text_wrap": True}
        )

        column_widths = [42, 105, 62, 171, 140, 67, 73]
        for col, width in enumerate(column_widths):
            worksheet.set_column(col, col, width / 7, align_format)

        worksheet.set_default_row(39)
        worksheet.set_margins(left=0.25, right=0.25, top=0.6, bottom=0.3)

        header_format = workbook.add_format(
            {
                "font_size": 15,
                "bold": True,
                "underline": True,
                "align": "center",
                "valign": "vcenter",
            }
        )
        worksheet.set_header(
            '&C&"Arial"&B{}&"'.format(self.configurations["examName"]),
            {"header_header_spacing": 0.3, "font": header_format},
        )

        for row_num, row_data in enumerate(df.values):
            for col_num, cell_value in enumerate(row_data):
                worksheet.write(row_num + 1, col_num, str(cell_value), align_format)

        worksheet.repeat_rows(0)

        max_row, max_col = df.shape
        column_settings = [{"header": column} for column in df.columns]
        worksheet.add_table(0, 0, max_row, max_col - 1, {"columns": column_settings})

    def extract_sheet(self, input_file: str, sheet_name: str, output_file: str) -> bool:
        """
        Extracts a single sheet from the input Excel file and saves it as a new file,
        preserving the original formatting.

        Args:
            input_file (str): Path to the input Excel file.
            sheet_name (str): Name of the sheet to extract.
            output_file (str): Path to save the extracted sheet.

        Returns:
            bool: True if extraction was successful, False otherwise.
        """
        try:
            new_wb = load_workbook(input_file)

            if sheet_name not in new_wb.sheetnames:
                self.log_message(
                    f"Sheet '{sheet_name}' not found in the workbook", logging.ERROR
                )
                return False

            for sheet in new_wb.sheetnames:
                if sheet != sheet_name:
                    new_wb.remove(new_wb[sheet])

            new_wb.save(output_file)

            self.log_message(
                f"Sheet '{sheet_name}' extracted and saved as '{output_file}'",
                logging.INFO,
            )
            return True

        except Exception as e:
            self.log_message(f"Error extracting sheet: {str(e)}", logging.ERROR)
            return False

    def get_sheet_names(self, excel_file: str) -> List[str]:
        """
        Returns a list of sheet names in the given Excel file.

        Args:
            excel_file (str): Path to the Excel file.

        Returns:
            List[str]: List of sheet names.
        """
        try:
            wb = pd.ExcelFile(excel_file)
            return wb.sheet_names
        except Exception as e:
            self.log_message(f"Error getting sheet names: {str(e)}", logging.ERROR)
            return []
